# parses through Summit Reforestation planting pay stubs, collects and formats
# data into excel file.

import datetime, os, PyPDF2, re, openpyxl, logging, sys
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
    BarChart,
    PieChart,
    series
)

logging.basicConfig(filename='log.log', filemode='w',
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    datefmt='%d-%b-%y %H:%M:%S',
                    level=logging.INFO)
logging.info('Start of log')

class Script_Instance:
    def __init__(self, mode: int = 0, empty: bool = 1, contract: bool = 1, 
                centage: bool = 1):
        """ Instance of script complete with arguments pertaining to optional
        data displays

        Args:
            mode (int): 0 = chart based on trees, 1 = chart based on pay
            empty (bool): 1 means populate dates with empty dates where no
            trees were planted (shows days off)
            contract (bool): display data based on contracts
            centage (bool): display data based on centage
        """
        self.mode = mode
        self.empty = empty
        self.contract = contract
        self.centage = centage

        self.run_script()

    def run_script(self):

        day_dict = {}
        PAYSTUB_DIR = 'summit_script/paystubs'
        season = Season()
        
        # loop through files in paystubs folder
        for paystub in os.listdir(PAYSTUB_DIR):
            logging.info(paystub)
            # scan file for data
            day_dict = scan_stub_reg(PAYSTUB_DIR, paystub, day_dict)
            #print(day_dict)

        # populate list from dictionary
        #tempdayList = day_dict.items()
        dayList = []
        for item in day_dict.items():
            dayList.append(item[1])

        dayList.sort()
        if self.empty == 1:
            dayList = add_empty_days(dayList)
        
        for item in dayList:
            #dayList.append(item[1])
            season.add_day(item)
            #print(item[1].date)
        
        
        #dayList.sort()
        #print(dayList)
        
        # Copy data into excel, for now: Date, Tree Total, Money Total
        rowCount = 1
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'Date'
        ws['B1'] = 'Total Trees'
        ws['C1'] = 'Pay'
        rowCount += 1
        
        for item in season.days:
            
            ws['A'+ str(rowCount)] = item.date.strftime('%d/%m/%Y')
            ws['B'+ str(rowCount)] = item.treestotal
            ws['C'+ str(rowCount)] = item.total
            
            rowCount += 1
            
        if self.contract == 1:
            # Copy Data into excel Contract, Trees Total, Pay Total, Avg Price
            ws['E35'] = 'BY CONTRACT'
            ws['E36'] = 'Contract'
            ws['F36'] = 'Trees'
            ws['G36'] = 'Pay'
            ws['H36'] = 'Avg. Price'
            
            rowCount_2 = 37
            block_data_list = sorted(season.dictcontracts.items())

            for item in block_data_list:
                ws['E' + str(rowCount_2)] = item[0]
                ws['F' + str(rowCount_2)] = item[1][0]
                ws['G' + str(rowCount_2)] = item[1][1]
                ws['H' + str(rowCount_2)] = item[1][1] / item[1][0]
                rowCount_2 += 1
            
        # Copy Data in excel Centage, Trees     Avg. Price
        if self.centage == 1:
            ws['J35'] = 'CENTAGE'
            ws['J36'] = 'Price'
            ws['K36'] = 'Trees'
            
            cent_list = sorted(season.centage.items())
            rowCount_2 = 37
            for item in cent_list:
                ws['J' + str(rowCount_2)] = item[0]
                ws['K' + str(rowCount_2)] = item[1]
                rowCount_2 += 1
                
            ws['M35'] = 'Avg. Price:'
            ws['N35'] = season.averageprice
            
            # Make pie chart
            
            pie = PieChart()
            labels = Reference(ws, min_col=10 ,min_row=37, max_row=rowCount_2)
            data = Reference(ws, min_col=11 , min_row=36, max_row=rowCount_2 - 1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = 'Centage Breakdown'
            ws.add_chart(pie, 'M37')
        
        # Copy Data in excel (total, best, worst, average) trees & pay
        # num of contacts, num of blocks
        
        ws['AE2'] = 'Trees'
        ws['AF2'] = 'Pay'
        
        ws['AD3'] = 'Total'
        ws['AE3'] = season.totaltrees
        ws['AF3'] = season.totalpay
        
        ws['AD4'] = 'Best'
        ws['AE4'] = season.bestdaytrees.treestotal
        ws['AF4'] = season.bestdaypay.total
        
        ws['AD5'] = 'Worst'
        ws['AE5'] = season.worstdaytrees.treestotal
        ws['AF5'] = season.worstdaypay.total
        
        ws['AD6'] = 'Average'
        ws['AE6'] = season.totaltrees / len(season.days)
        ws['AF6'] = season.totalpay / len(season.days)
        
        ws['AD8'] = '# of Contracts'
        ws['AE8'] = len(season.dictcontracts)
        
        ws['AD9'] = '# of Blocks'
        ws['AE9'] = len(season.dictblocks)
        
        # Create Tree Chart
        c1 = BarChart()
        c1.type = 'col'
        c1.style = 10 # aka theme colours
        c1.title = 'Pay by Day'
        c1.y_axis.title = 'Gross Pay (CAD)'
        c1.x_axis.title = 'Date'
        
        data = Reference(ws, min_col=3, min_row=2, max_row=rowCount-1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=rowCount-1)
        
        c1.add_data(data, titles_from_data=False)
        c1.set_categories(cats)
        
        c1.width = 40
        c1.height = 15
        

        ws.add_chart(c1, 'E3')
        
        wb.save('summit_script/output/paystub.xlsx')

class Day:
    """ Stores info about specific days in planting season
    
    Args:
        date (datetime.date): the date
        trees (list[tuple]): tuple format (# of trees, price per tree, 
        total pay)
        treestotal (int): total amount of trees for Day
        total (float): total pay for Day
    """
    def __init__(self, date: datetime.date) -> None:
        
        self.date = date
        self.treedata = []
        # treedata format (contract, block, trees, inclusive, base, net, stat,
        # vac, total)
        
        self.treestotal = 0
        self.total = 0
        self.empty = 1
        
    def addTreeData(self, contract: str, block: str, trees: int,
                    inclusive: float, base: float, net: float, 
                    stat: float, vac: float, total: float) -> None:
        
        self.empty = 0
        treeline = (contract, block, trees, inclusive, base, net, stat, vac, 
                    total)
        self.treedata.append(treeline)
        self.treestotal += trees
        self.total += total
        
    def __eq__(self, other):
        return self.date == other.date
    
    def __lt__(self, other):
        return self.date < other.date
    
    def __gt__(self, other):
        return self.date > other.date
    
    def __ge__(self, other):
        return self.date >= other.date
    
    def __le__(self, other):
        return self.date <= other.date
    
class Season:
    
    def __init__(self):
        self.days = []
        
        # num contracts format =    017 : (1000, $130)
        self.dictcontracts = {}
        # num blocks format     JB21 : (10000,$1300)
        self.dictblocks = {}
        # total num trees
        self.totaltrees = 0
        # average trees
        self.averagetrees = 0
        # total gross pay
        self.totalpay = 0
        # average gross pay
        self.averagepay = 0
        # centage list format = 0.13 : 89000     centage : total trees
        self.centage = {}
        # average price
        self.averageprice = 0
        # best day trees
        self.bestdaytrees = None
        # best day pay
        self.bestdaypay = None
        # worst day trees
        self.worstdaytrees = None
        # worst day pay
        self.worstdaypay = None
        
    def add_day(self, newday: Day):
        self.days.append(newday)
        self.days.sort
        
        treedata = newday.treedata
        
        for line in treedata:
            #print(line[0])
            if self.dictcontracts.get(line[0]) is None:
                self.dictcontracts[line[0]] = (line[2],line[8])
            else:
                temptuple = self.dictcontracts[line[0]]
                self.dictcontracts[line[0]] = (temptuple[0]+line[2],
                                            temptuple[1]+line[8])
                
            if self.dictblocks.get(line[1]) is None:
                self.dictblocks[line[1]] = (line[2], line[8])
            else:
                temptuple = self.dictblocks[line[1]]
                self.dictblocks[line[1]] = (temptuple[0]+line[2],
                                            temptuple[1]+line[8])
                
            self.totaltrees += line[2]
            self.totalpay += line[8]
            
            if line[3] not in self.centage:
                self.centage[line[3]] = line[2]
            else:
                self.centage[line[3]] += line[2]
                
        if self.bestdaytrees is None:
            self.bestdaytrees = newday
        elif self.bestdaytrees.treestotal < newday.treestotal:
            self.bestdaytrees = newday
            
        if self.bestdaypay is None:
            self.bestdaypay = newday
        elif self.bestdaypay.total < newday.total:
            self.bestdaypay = newday
            
        if self.worstdaytrees is None:
            self.worstdaytrees = newday
        elif self.worstdaytrees.treestotal > newday.treestotal:
            self.worstdaytrees = newday
            
        if self.worstdaypay is None:
            self.worstdaypay = newday
        elif self.worstdaypay.total > newday.total:
            self.worstdaypay = newday
                    
        self.averagetrees = self.totaltrees / len(self.days)
        self.averagepay = self.totalpay / len(self.days)
        self.averageprice = self.totalpay / self.totaltrees
                
            


def get_date_difference(startdate, enddate):
    """

    AWHAT IS THIS?
    
    """
    diff = enddate - startdate
    return diff.days

def scan_stub(dir: str, filename: str, daydict: dict) -> dict:
    days = daydict
    
    MONTHS = {
        'Jan': 0,
        'Feb': 1,
        'Mar': 2,
        'Apr': 3,
        'May': 4,
        'Jun': 5,
        'Jul': 6,
        'Aug': 7,
        'Sep': 8,
        'Oct': 9,
        'Nov': 10,
        'Dec': 11
    }


    # open file
    pdfFileObject = open(dir + '/' + filename, 'rb')
    pdfReader = PyPDF2.PdfReader(pdfFileObject)
    
    flag = 1
    count = 0
    while flag:
    # loop through pages in paystub file
        
        try:
            pageObj = pdfReader.pages[count]
            pageText = pageObj.extract_text()
            pageLines = pageText.split('\n')
            count += 1
    
            for line in pageLines:
                tempLine = line.split(' ')
                
                if len(tempLine) == 10:
                    splitDate = tempLine[0].split('-')
                    datee = datetime.date(int('20' + splitDate[2]), 
                                        MONTHS[splitDate[1]],
                                        int(splitDate[0]))
                    
                    trees = int(tempLine[3].replace(',',''))
                    price = float(tempLine[8])
                    
                    
                    if tempLine[0] in days:
                        days[tempLine[0]].addTreeData(trees,price)
                    else:
                        days[tempLine[0]] = Day(datee, (trees,price))
        except:
            flag = 0
            
    return days

def scan_stub_reg(dir: str, filename: str, daydict: dict) -> dict:
    days = daydict
    
    MONTHS = {
        'Jan': 1,
        'Feb': 2,
        'Mar': 3,
        'Apr': 4,
        'May': 5,
        'Jun': 6,
        'Jul': 7,
        'Aug': 8,
        'Sep': 9,
        'Oct': 10,
        'Nov': 11,
        'Dec': 12
    }
    
    #regex = r"""(\d\d-\w{3}-\d\d)\s+(\d{3})\s+([a-zA-Z0-9-]+)\s+([0-9,]+)\s+([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)"""
    pattern = re.compile(r'(\d\d-\w{3}-\d\d)\s+(\d{3})\s+([a-zA-Z0-9-]+)\s+([0-9,]+)\s+([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)')
    
    # (1)Date, (2)Contract, (3)Block, (4)Trees, (5)Inclusive, (6)Base,
    # (7)Net, (8)Stat, (9)Vac, (10)Total   <- Regex group key
        
    # open file
    pdfFileObject = open(dir + '/' + filename, 'rb')
    pdfReader = PyPDF2.PdfReader(pdfFileObject)
    
    flag = 1
    count = 0
    while flag:
    # loop through pages in paystub file
        
        try:
            pageObj = pdfReader.pages[count]
            pageText = pageObj.extract_text()
            matches = pattern.finditer(pageText)
            logging.info(matches)
            for match in matches:
                logging.info(f'Match found: {match}')
                date_list = match.group(1).split('-')
                # e.g. 28-May-23
                date_dt = datetime.date(year = int('20'+date_list[2]),
                                        month = MONTHS[date_list[1]],
                                        day = int(date_list[0]))
                contract_str = match.group(2)
                block_str = match.group(3)
                trees_int = int(match.group(4).replace(',',''))
                inclusive_float = float(match.group(9))
                base_float = float(match.group(5))
                net_float = float(match.group(6))
                stat_float = float(match.group(7))
                vac_float = float(match.group(8))
                total_float = float(match.group(10))
                
                
                if match.group(1) in days:
                    days[match.group(1)].addTreeData(contract_str, block_str,
                                                    trees_int, inclusive_float,
                                                    base_float, net_float,
                                                    stat_float, vac_float,
                                                    total_float)
                else:
                    days[match.group(1)] = Day(date_dt)
                    days[match.group(1)].addTreeData(contract_str, block_str,
                                                    trees_int, inclusive_float,
                                                    base_float, net_float,
                                                    stat_float, vac_float,
                                                    total_float)
            count += 1
        except:
            flag = 0
            
    return days

def add_empty_days(day_list: list[Day]) -> list[Day]:
    listy = day_list
    listy_updated = []
    if len(listy) > 1:
        prevDate = listy[0].date
        listy_updated.append(listy[0])
        for i in range(1, len(listy)):
            currDate = listy[i].date
            diff = get_date_difference(prevDate, currDate)
            
            if diff == 0:
                logging.warning('duplicate dates in dayList')
            elif diff > 1:
                for j in range(1, diff):
                    fixDate = prevDate + datetime.timedelta(days = j)
                    tempDay = Day(fixDate)
                    listy_updated.append(tempDay)
            
            listy_updated.append(listy[i])
            prevDate = listy[i].date
    
    listy_updated.sort()
    #for item in listy_updated:
        #print(item.date)
    return listy_updated

def arg_checker(arg_list: list) -> bool:
    if len(arg_list) == 1:
        return True
    elif len(arg_list) == 5:
        flag = 0
        if arg_list[1] != '0' and arg_list[1] != '1':
            print("Arg 1 must be 0 or 1")
            flag = 1
        if arg_list[2] != '0' and arg_list[2] != '1':
            print("Arg 2 must be 0 or 1")
            flag = 1
        if arg_list[3] != '0' and arg_list[3] != '1':
            print("Arg 3 must be 0 or 1")
            flag = 1
        if arg_list[4] != '0' and arg_list[4] != '1':
            print("Arg 3 must be 0 or 1")
            flag = 1

        if flag == 1:
            return False
        else: return True
    else:
        print("Incorrect amount of arguments, 0 or 4 accepted.")
        return False


if __name__ == '__main__':
    
    if arg_checker(sys.argv):
        if len(sys.argv) == 1:
            script = Script_Instance()
        else:
            script = Script_Instance(int(sys.argv[1]), int(sys.argv[2]), 
                                    int(sys.argv[3]), int(sys.argv[4]))
    
    
